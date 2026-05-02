from fastapi import APIRouter, Depends, UploadFile, File, Query
from sqlalchemy.ext.asyncio import AsyncSession
from app.db.session import get_db
from app.services.dashboard import DashboardService
from app.services.import_service import ImportService
from app.core.deps import get_current_active_user
from app.models.user import User

dashboard_router = APIRouter(prefix="/dashboard", tags=["dashboard"])


@dashboard_router.get("/summary")
async def get_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DashboardService(db)
    return await service.get_summary()


@dashboard_router.get("/payments-by-day")
async def get_payments_by_day(
    days: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DashboardService(db)
    return await service.get_payments_by_day(days)


@dashboard_router.get("/top-managers")
async def get_top_managers(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DashboardService(db)
    return await service.get_top_managers()


# Import
import_router = APIRouter(prefix="/import", tags=["import"])


@import_router.post("/excel")
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    contents = await file.read()
    service = ImportService(db)
    log = await service.import_excel(contents, file.filename or "upload.xlsx", current_user.id)
    return {
        "id": log.id,
        "filename": log.filename,
        "total_rows": log.total_rows,
        "success_rows": log.success_rows,
        "error_rows": log.error_rows,
        "status": log.status,
        "errors": log.errors,
    }


@dashboard_router.get("/overdue-promises")
async def overdue_promises(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    from sqlalchemy import select
    from app.models.operations import Promise
    from datetime import date
    result = await db.execute(
        select(Promise)
        .where(Promise.status == "active", Promise.promise_date < date.today())
        .order_by(Promise.promise_date)
        .limit(50)
    )
    promises = result.scalars().all()
    return {
        "count": len(promises),
        "items": [
            {
                "id": p.id,
                "contract_id": p.contract_id,
                "amount": float(p.amount),
                "promise_date": str(p.promise_date),
                "days_overdue": (date.today() - p.promise_date).days,
            }
            for p in promises
        ]
    }


@dashboard_router.get("/calls-history")
async def calls_history(
    manager_id: int = None,
    date_from: str = None,
    date_to: str = None,
    page: int = 1,
    page_size: int = 50,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    from sqlalchemy import select, func, and_
    from app.models.operations import CallLog
    from app.models.user import User as UserModel
    from datetime import datetime

    # Managers see only their calls
    if current_user.role.upper() == "MANAGER":
        manager_id = current_user.id

    q = select(CallLog)
    filters = []
    if manager_id:
        filters.append(CallLog.manager_id == manager_id)
    if date_from:
        filters.append(CallLog.called_at >= datetime.fromisoformat(date_from))
    if date_to:
        filters.append(CallLog.called_at <= datetime.fromisoformat(date_to + "T23:59:59"))
    if filters:
        q = q.where(and_(*filters))

    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar_one()
    calls = (await db.execute(q.order_by(CallLog.called_at.desc()).offset((page-1)*page_size).limit(page_size))).scalars().all()

    return {
        "total": total,
        "page": page,
        "items": [
            {
                "id": c.id,
                "contract_id": c.contract_id,
                "manager_id": c.manager_id,
                "result": c.result,
                "duration_seconds": c.duration_seconds,
                "notes": c.notes,
                "called_at": str(c.called_at),
            }
            for c in calls
        ]
    }


@dashboard_router.get("/export-debtors")
async def export_debtors(
    contract_status: str = None,
    manager_id: int = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export debtors to Excel."""
    from fastapi.responses import StreamingResponse
    from sqlalchemy import select, and_
    from app.models.debtor import Debtor
    from app.models.contract import Contract
    from app.models.operations import Assignment
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    # Build query
    q = (
        select(Debtor, Contract)
        .join(Contract, Contract.debtor_id == Debtor.id)
    )
    filters = []
    if contract_status:
        filters.append(Contract.status == contract_status)
    if manager_id:
        assigned = select(Assignment.contract_id).where(
            and_(Assignment.manager_id == manager_id, Assignment.is_active == True)
        )
        filters.append(Contract.id.in_(assigned))
    if current_user.role.upper() == "MANAGER":
        assigned = select(Assignment.contract_id).where(
            and_(Assignment.manager_id == current_user.id, Assignment.is_active == True)
        )
        filters.append(Contract.id.in_(assigned))
    if filters:
        q = q.where(and_(*filters))

    rows = (await db.execute(q.limit(10000))).all()

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Должники"

    headers = ["ИИН", "ФИО", "Телефон", "Адрес", "Номер договора", "Статус", "Итого долг", "ОД", "Вознаграждение", "Штрафы"]
    header_fill = PatternFill("solid", fgColor="366092")
    header_font = Font(bold=True, color="FFFFFF")

    STATUS_LABELS = {"active": "Досудебный", "litigation": "Судебный", "closed": "Закрыт", "written_off": "Списан"}

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, (debtor, contract) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=debtor.iin)
        ws.cell(row=i, column=2, value=debtor.full_name)
        ws.cell(row=i, column=3, value=debtor.phone_primary or "")
        ws.cell(row=i, column=4, value=debtor.address or "")
        ws.cell(row=i, column=5, value=contract.contract_number)
        ws.cell(row=i, column=6, value=STATUS_LABELS.get(contract.status, contract.status))
        ws.cell(row=i, column=7, value=float(contract.total_debt))
        ws.cell(row=i, column=8, value=float(contract.principal_debt))
        ws.cell(row=i, column=9, value=float(contract.interest_debt))
        ws.cell(row=i, column=10, value=float(contract.penalty_debt))

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=debtors_export.xlsx"}
    )
