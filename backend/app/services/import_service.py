"""
Импорт данных из файлов Hall (Excel).

Листы:
  - Остатки      — должники + договоры
  - Платежи      (11 кол.) — история платежей по договорам
  - Отзывы       (7 кол.)  — договоры, отозванные из агентства
  - Новые_Номера (3 кол.)  — дополнительные телефоны должников
"""
import io
from datetime import datetime, date as date_type
from decimal import Decimal, InvalidOperation
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models.operations import ImportLog, Payment
from app.models.debtor import Debtor
from app.models.contract import Contract
from app.repositories.debtor import DebtorRepository
from app.repositories.contract import ContractRepository

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


STATUS_MAP = {
    "досуд.": "active",
    "исполнад": "litigation",
    "суд": "litigation",
    "закрыт": "closed",
    "списан": "written_off",
}


class ImportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.debtor_repo = DebtorRepository(db)
        self.contract_repo = ContractRepository(db)

    # ─── Точка входа ──────────────────────────────────────────────────────────

    async def import_excel(self, file_bytes: bytes, filename: str, uploaded_by_id: int) -> ImportLog:
        log = ImportLog(filename=filename, uploaded_by_id=uploaded_by_id, status="processing")
        self.db.add(log)
        await self.db.flush()

        errors: list[dict] = []
        success = 0
        total = 0

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

            # 1. Остатки — должники + договоры
            if "Остатки" in wb.sheetnames:
                s, t, e = await self._import_ostatok(wb["Остатки"])
                success += s; total += t; errors.extend(e)

            # 2. Платежи — история платежей (только КЭШ и Внутреннее ПТ — не банковские)
            if "Платежи" in wb.sheetnames:
                s, t, e = await self._import_payments(wb["Платежи"], uploaded_by_id)
                success += s; total += t; errors.extend(e)

            # 3. Новые_Номера — дополнительные телефоны
            if "Новые_Номера" in wb.sheetnames:
                s, t, e = await self._import_new_phones(wb["Новые_Номера"])
                success += s; total += t; errors.extend(e)

            # 4. Отзывы — закрыть договоры
            if "Отзывы" in wb.sheetnames:
                s, t, e = await self._import_recalls(wb["Отзывы"])
                success += s; total += t; errors.extend(e)

        except Exception as e:
            errors.append({"row": 0, "error": f"Ошибка чтения файла: {e}"})

        log.total_rows = total
        log.success_rows = success
        log.error_rows = len(errors)
        log.errors = errors[:200]
        log.status = "done" if not errors else "done_with_errors"
        log.finished_at = datetime.utcnow()
        await self.db.flush()
        return log

    # ─── Лист: Остатки ────────────────────────────────────────────────────────

    async def _import_ostatok(self, ws) -> tuple[int, int, list]:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)
        success = 0
        errors = []

        for idx, row in enumerate(rows):
            row_num = idx + 2
            try:
                iin = self._str(row, 11)   # col 12
                if not iin:
                    raise ValueError("ИИН отсутствует")
                iin = "".join(c for c in iin if c.isdigit())
                if not iin:
                    raise ValueError("ИИН не содержит цифр")

                contract_number = self._str(row, 8)   # col 9
                if not contract_number:
                    raise ValueError("Номер договора отсутствует")

                full_name    = self._str(row, 12) or "Неизвестно"
                phone_raw    = self._str(row, 51) or ""   # col 52
                address      = self._str(row, 52)          # col 53
                special_notes = self._str(row, 53)         # col 54

                # Должник
                debtor = await self.debtor_repo.get_by_iin(iin)
                if not debtor:
                    debtor = Debtor(
                        iin=iin,
                        full_name=full_name,
                        phone_primary=self._first_phone(phone_raw),
                        address=address,
                        notes=special_notes,
                    )
                    self.db.add(debtor)
                    await self.db.flush()
                else:
                    # Обновляем телефон/адрес если они пустые
                    if not debtor.phone_primary and phone_raw:
                        debtor.phone_primary = self._first_phone(phone_raw)
                    if not debtor.address and address:
                        debtor.address = address

                # Договор
                contract = await self.contract_repo.get_by_number(contract_number)

                total_debt     = self._dec(row, 14)  # col 15
                principal_debt = self._dec(row, 15)  # col 16
                interest_debt  = self._dec(row, 16)  # col 17
                penalty_debt   = self._dec(row, 17)  # col 18
                issue_date     = self._date(row, 9)  # col 10
                close_date     = self._date(row, 10) # col 11

                status_raw = (self._str(row, 26) or "").strip().lower()
                status = STATUS_MAP.get(status_raw, "active")

                creditor  = self._str(row, 3) or "Hall"   # col 4 Филиал
                product   = self._str(row, 5) or ""        # col 6 Продукт
                registry  = self._str(row, 24) or ""       # col 25 Номер реестра

                if not contract:
                    contract = Contract(
                        debtor_id=debtor.id,
                        contract_number=contract_number,
                        original_creditor=creditor,
                        product_type=product,
                        principal_debt=principal_debt,
                        interest_debt=interest_debt,
                        penalty_debt=penalty_debt,
                        total_debt=total_debt,
                        currency="KZT",
                        issue_date=issue_date,
                        overdue_date=close_date,
                        status=status,
                        notes=registry,
                    )
                    self.db.add(contract)
                else:
                    # Обновляем долг
                    contract.total_debt     = total_debt
                    contract.principal_debt = principal_debt
                    contract.interest_debt  = interest_debt
                    contract.penalty_debt   = penalty_debt
                    contract.status         = status

                await self.db.flush()
                success += 1

            except Exception as e:
                errors.append({"sheet": "Остатки", "row": row_num, "error": str(e)})

        return success, total, errors

    # ─── Лист: Платежи ────────────────────────────────────────────────────────

    async def _import_payments(self, ws, registered_by_id: int) -> tuple[int, int, list]:
        """
        Колонки (0-based index):
         0 агентство    3 Сумма погашения   7  Номер договора
         1 Дата отчета  4 Внешнее ПТ        9  ИИН/БИН
         2 ID АБИС      5 Внутреннее ПТ    10  Клиент
                        6 КЭШ

        Логика:
        - Внешнее ПТ (col 4) = bank — ПРОПУСКАЕМ (банковские, не нужны)
        - Внутреннее ПТ (col 5) = card — ИМПОРТИРУЕМ
        - КЭШ (col 6) = cash — ИМПОРТИРУЕМ
        """
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)
        success = 0
        errors = []

        for idx, row in enumerate(rows):
            row_num = idx + 2
            try:
                contract_number = self._str(row, 7)   # col 8
                payment_date    = self._date(row, 1)  # col 2

                # Суммы по типам
                bank_amount     = self._dec(row, 4)   # Внешнее ПТ
                card_amount     = self._dec(row, 5)   # Внутреннее ПТ
                cash_amount     = self._dec(row, 6)   # КЭШ

                if not contract_number:
                    raise ValueError("Номер договора отсутствует")

                contract = await self.contract_repo.get_by_number(contract_number)
                if not contract:
                    success += 1
                    continue

                pay_date = payment_date or date_type.today()

                # Импортируем только НЕ банковские платежи
                # КЭШ платеж
                if cash_amount > 0:
                    exists = await self.db.execute(
                        select(Payment).where(
                            Payment.contract_id == contract.id,
                            Payment.amount == cash_amount,
                            Payment.payment_date == pay_date,
                            Payment.source == "cash",
                        )
                    )
                    if not exists.scalar_one_or_none():
                        self.db.add(Payment(
                            contract_id=contract.id,
                            amount=cash_amount,
                            payment_date=pay_date,
                            source="cash",
                            reference="Hall-import",
                            registered_by_id=registered_by_id,
                        ))
                        await self.db.flush()

                # Внутреннее ПТ (card)
                if card_amount > 0:
                    exists = await self.db.execute(
                        select(Payment).where(
                            Payment.contract_id == contract.id,
                            Payment.amount == card_amount,
                            Payment.payment_date == pay_date,
                            Payment.source == "card",
                        )
                    )
                    if not exists.scalar_one_or_none():
                        self.db.add(Payment(
                            contract_id=contract.id,
                            amount=card_amount,
                            payment_date=pay_date,
                            source="card",
                            reference="Hall-import",
                            registered_by_id=registered_by_id,
                        ))
                        await self.db.flush()

                success += 1

            except Exception as e:
                errors.append({"sheet": "Платежи", "row": row_num, "error": str(e)})

        return success, total, errors

    # ─── Лист: Новые_Номера ───────────────────────────────────────────────────

    async def _import_new_phones(self, ws) -> tuple[int, int, list]:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)
        success = 0
        errors = []

        for idx, row in enumerate(rows):
            row_num = idx + 2
            try:
                iin = self._str(row, 1)
                if not iin:
                    raise ValueError("ИИН отсутствует")
                iin = "".join(c for c in iin if c.isdigit())

                phone_raw = self._str(row, 2) or ""
                phone = self._first_phone(phone_raw)
                if not phone:
                    success += 1
                    continue

                debtor = await self.debtor_repo.get_by_iin(iin)
                if debtor:
                    if not debtor.phone_secondary and debtor.phone_primary != phone:
                        debtor.phone_secondary = phone
                    await self.db.flush()

                success += 1
            except Exception as e:
                errors.append({"sheet": "Новые_Номера", "row": row_num, "error": str(e)})

        return success, total, errors

    # ─── Лист: Отзывы ────────────────────────────────────────────────────────

    async def _import_recalls(self, ws) -> tuple[int, int, list]:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)
        success = 0
        errors = []

        for idx, row in enumerate(rows):
            row_num = idx + 2
            try:
                dbz_raw = self._str(row, 1)  # col 2 ДБЗ
                if not dbz_raw:
                    success += 1
                    continue

                # Номер договора в ДБЗ может быть с суффиксом /500
                contract_number = dbz_raw.split("/")[0] if "/" in dbz_raw else dbz_raw

                contract = await self.contract_repo.get_by_number(contract_number)
                if contract and contract.status not in ("closed", "written_off"):
                    contract.status = "closed"
                    await self.db.flush()

                success += 1
            except Exception as e:
                errors.append({"sheet": "Отзывы", "row": row_num, "error": str(e)})

        return success, total, errors

    # ─── Вспомогательные методы ───────────────────────────────────────────────

    def _str(self, row, idx: int) -> str | None:
        if idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        return str(v).strip() or None

    def _dec(self, row, idx: int) -> Decimal:
        v = self._str(row, idx)
        if not v:
            return Decimal("0")
        v = v.replace(" ", "").replace(",", ".")
        try:
            return Decimal(v)
        except InvalidOperation:
            return Decimal("0")

    def _date(self, row, idx: int):
        if idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        if isinstance(v, (datetime,)):
            return v.date()
        if isinstance(v, date_type):
            return v
        try:
            return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
        except Exception:
            return None

    def _first_phone(self, raw: str) -> str | None:
        if not raw:
            return None
        import re
        phones = re.findall(r"[+7][7\d]\d{9}", raw.replace(" ", "").replace("-", ""))
        if phones:
            return phones[0]
        digits = re.findall(r"\d{10,11}", raw.replace(" ", ""))
        if digits:
            d = digits[0]
            if len(d) == 10:
                return "+" + "7" + d
            if len(d) == 11 and d.startswith("8"):
                return "+7" + d[1:]
            return "+" + d
        return None
